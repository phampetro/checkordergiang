import os
import sys
import openpyxl
import warnings
from pathlib import Path
from datetime import datetime

# Tắt warning openpyxl về default style
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

class ExcelProcessor:
    def __init__(self):
        self.base_path = Path(__file__).parent
        self.output_dir = self.base_path / "output"
        
    def get_daily_directory(self):
        """
        Lấy thư mục output theo ngày hiện tại (format: DDMMYYYY)
        """
        today = datetime.now().strftime("%d%m%Y")
        daily_dir = self.output_dir / today
        
        if daily_dir.exists():
            return daily_dir
        else:
            print(f"❌ Không tìm thấy thư mục: {daily_dir}")
            return None
    
    def process_excel_files(self):
        """
        Xử lý tất cả file Excel trong thư mục ngày hiện tại
        """
        daily_dir = self.get_daily_directory()
        if not daily_dir:
            return False
        
        # Tìm tất cả file Excel
        excel_files = list(daily_dir.glob("*.xlsx"))
        if not excel_files:
            print(f"❌ Không tìm thấy file Excel nào trong: {daily_dir}")
            return False
        
        print(f"📁 Thư mục xử lý: {daily_dir}")
        print(f"📄 Tìm thấy {len(excel_files)} file Excel")
        print("─" * 50)
        
        success_count = 0
        for i, excel_file in enumerate(excel_files, 1):
            print(f"🔄 Xử lý file {i}/{len(excel_files)}: {excel_file.name}")
            
            if self.process_single_excel(excel_file):
                success_count += 1
                print(f"   ✅ Hoàn thành: {excel_file.name}")
            else:
                print(f"   ❌ Thất bại: {excel_file.name}")
            
            print()
        
        print("─" * 50)
        print(f"📊 Kết quả: {success_count}/{len(excel_files)} file được xử lý thành công")
        
        return success_count > 0
    
    def process_single_excel(self, excel_file):
        """
        Xử lý một file Excel theo từng bước tuần tự:
        5 dòng đầu tiêu đề
        B1: Ẩn từ dòng 1 đến dòng 3
        B2: Ẩn dòng có cột A rỗng
        B3: Ẩn dòng có cột B rỗng
        B4: Ẩn dòng có cột D rỗng AND cột C <> ""
        B4: Xóa dữ liệu của các dòng có cột C rỗng, xóa từ K trở đi
        B5: Ẩn các dòng K có chứa nội dung "NPP bán"
        B6: Ẩn dòng có cột Q > 0 (giữ lại dòng rỗng và 0)
        B7: Kiểm tra cột Q nếu có 2 dòng rỗng liên tiếp thì ẩn dòng thứ 2
        B8: Ẩn cột S trở đi, cột A đến F, cột M và N
        B9: Cố định xem được tiêu đề
        """
        try:
            # Mở file Excel
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            row_count = ws.max_row
            col_count = ws.max_column
            
            print(f"   📊 Kích thước: {row_count} dòng × {col_count} cột")
            
            # B1: Ẩn từ dòng 1 đến dòng 3
            for row_num in range(1, 4):  # Dòng 1, 2, 3
                ws.row_dimensions[row_num].hidden = True
            print(f"   🙈 B1: Đã ẩn dòng 1-3")
            
            # B2: Ẩn dòng có cột A rỗng (từ dòng 6 trở đi, bỏ qua tiêu đề dòng 4-5)
            hidden_count_a = 0
            for row_num in range(6, row_count + 1):
                cell_a = ws.cell(row_num, 1)
                if cell_a.value is None or str(cell_a.value).strip() == "":
                    ws.row_dimensions[row_num].hidden = True
                    hidden_count_a += 1
            print(f"   🙈 B2: Đã ẩn {hidden_count_a} dòng có cột A rỗng")
            
            # B3: Ẩn dòng có cột B rỗng (chỉ xét dòng chưa bị ẩn)
            hidden_count_b = 0
            for row_num in range(6, row_count + 1):
                if not ws.row_dimensions[row_num].hidden:
                    cell_b = ws.cell(row_num, 2)
                    if cell_b.value is None or str(cell_b.value).strip() == "":
                        ws.row_dimensions[row_num].hidden = True
                        hidden_count_b += 1
            print(f"   🙈 B3: Đã ẩn {hidden_count_b} dòng có cột B rỗng")
            
            # B4: Ẩn dòng có cột D rỗng AND cột C <> ""
            hidden_count_d = 0
            for row_num in range(6, row_count + 1):
                if not ws.row_dimensions[row_num].hidden:
                    cell_c = ws.cell(row_num, 3)
                    cell_d = ws.cell(row_num, 4)
                    
                    c_not_empty = (cell_c.value is not None and str(cell_c.value).strip() != "")
                    d_empty = (cell_d.value is None or str(cell_d.value).strip() == "")
                    
                    if d_empty and c_not_empty:
                        ws.row_dimensions[row_num].hidden = True
                        hidden_count_d += 1
            print(f"   🙈 B4: Đã ẩn {hidden_count_d} dòng có cột D rỗng AND C khác rỗng")
            
            # B4: Xóa dữ liệu của các dòng có cột C rỗng, xóa từ K trở đi
            cleared_count_c = 0
            for row_num in range(6, row_count + 1):
                cell_c = ws.cell(row_num, 3)
                if cell_c.value is None or str(cell_c.value).strip() == "":
                    # Xóa dữ liệu từ cột K (11) trở đi
                    if col_count >= 11:
                        for col_num in range(11, col_count + 1):
                            try:
                                ws.cell(row_num, col_num).value = None
                            except AttributeError:
                                pass  # Bỏ qua MergedCell
                        cleared_count_c += 1
            print(f"   🗑️ B4: Đã xóa dữ liệu từ cột K trở đi trên {cleared_count_c} dòng có cột C rỗng")
            
            # B5: Ẩn các dòng K có chứa nội dung "NPP bán"
            hidden_count_k = 0
            for row_num in range(6, row_count + 1):
                if not ws.row_dimensions[row_num].hidden:
                    cell_k = ws.cell(row_num, 11)  # Cột K
                    if cell_k.value is not None and "NPP bán" in str(cell_k.value):
                        ws.row_dimensions[row_num].hidden = True
                        hidden_count_k += 1
            print(f"   🙈 B5: Đã ẩn {hidden_count_k} dòng có cột K chứa 'NPP bán'")
            
            # B6: Ẩn dòng có cột Q > 0 (giữ lại dòng rỗng và 0)
            hidden_count_q = 0
            for row_num in range(6, row_count + 1):
                if not ws.row_dimensions[row_num].hidden:
                    cell_q = ws.cell(row_num, 17)  # Cột Q
                    if cell_q.value is not None:
                        try:
                            q_value = float(cell_q.value)
                            if q_value > 0:
                                ws.row_dimensions[row_num].hidden = True
                                hidden_count_q += 1
                        except (ValueError, TypeError):
                            pass
            print(f"   🙈 B6: Đã ẩn {hidden_count_q} dòng có cột Q > 0")
            
            # B7: Kiểm tra cột Q nếu có 2 dòng rỗng liên tiếp thì ẩn dòng thứ 2
            hidden_count_q2 = 0
            prev_row_q_empty = False
            for row_num in range(6, row_count + 1):
                if not ws.row_dimensions[row_num].hidden:
                    cell_q = ws.cell(row_num, 17)  # Cột Q
                    current_row_q_empty = (cell_q.value is None or str(cell_q.value).strip() == "")
                    
                    if prev_row_q_empty and current_row_q_empty:
                        ws.row_dimensions[row_num].hidden = True
                        hidden_count_q2 += 1
                    
                    prev_row_q_empty = current_row_q_empty
            print(f"   🙈 B7: Đã ẩn {hidden_count_q2} dòng thứ 2 trong các cặp dòng Q rỗng liên tiếp")
            
            # B8: Ẩn cột S trở đi, cột A đến F, cột M và N
            hidden_cols = self.hide_unwanted_columns(ws)
            print(f"   👁️ B8: Đã ẩn {hidden_cols} cột: A-F, M, N và từ cột S trở đi")
            
            # B9: Cố định xem được tiêu đề
            ws.freeze_panes = "A6"  # Cố định dòng 4-5 (tiêu đề)
            print(f"   📌 B9: Đã cố định tiêu đề (freeze panes tại A6)")
            
            # Lưu file
            wb.save(excel_file)
            wb.close()
            
            return True
            
        except Exception as e:
            print(f"   ❌ Lỗi xử lý: {str(e)}")
            return False
    
    def hide_unwanted_columns(self, ws):
        """
        Ẩn cột A đến F, cột M, N và từ cột S trở đi.
        Trả về số cột đã ẩn.
        """
        try:
            total_cols = ws.max_column
            hidden_count = 0
            
            # Ẩn cột A đến F (cột 1 đến 6)
            for col_num in range(1, 7):  # Cột A=1 đến F=6
                if col_num <= total_cols:
                    ws.column_dimensions[ws.cell(1, col_num).column_letter].hidden = True
                    hidden_count += 1
            
            # Ẩn cột M (cột 13)
            if total_cols >= 13:
                ws.column_dimensions[ws.cell(1, 13).column_letter].hidden = True
                hidden_count += 1
            
            # Ẩn cột N (cột 14)
            if total_cols >= 14:
                ws.column_dimensions[ws.cell(1, 14).column_letter].hidden = True
                hidden_count += 1
            
            # Ẩn từ cột S trở đi (cột 19 trở đi)
            for col_num in range(19, total_cols + 1):  # Từ S=19 đến cuối
                ws.column_dimensions[ws.cell(1, col_num).column_letter].hidden = True
                hidden_count += 1
            
            return hidden_count
            
        except Exception as e:
            print(f"   ❌ Lỗi ẩn cột: {str(e)}")
            return 0
    
    def list_files_in_daily_directory(self):
        """
        Liệt kê các file trong thư mục ngày hiện tại
        """
        daily_dir = self.get_daily_directory()
        if not daily_dir:
            return
        
        excel_files = list(daily_dir.glob("*.xlsx"))
        
        print(f"📁 Thư mục: {daily_dir}")
        print(f"📄 Số file Excel: {len(excel_files)}")
        print("─" * 50)
        
        for i, excel_file in enumerate(excel_files, 1):
            file_size = excel_file.stat().st_size
            print(f"{i:2}. {excel_file.name} ({file_size:,} bytes)")
        
        if not excel_files:
            print("   (Không có file Excel nào)")

def main():
    """
    Hàm main để test chức năng xử lý Excel
    """
    processor = ExcelProcessor()
    
    print("🔧 EXCEL PROCESSOR")
    print("=" * 50)
    
    # Liệt kê file trước khi xử lý
    print("📋 DANH SÁCH FILE TRƯỚC KHI XỬ LÝ:")
    processor.list_files_in_daily_directory()
    print()
    
    # Xử lý các file Excel
    print("🚀 BẮT ĐẦU XỬ LÝ:")
    success = processor.process_excel_files()
    
    if success:
        print("✅ Hoàn thành xử lý Excel!")
    else:
        print("❌ Có lỗi trong quá trình xử lý!")

if __name__ == "__main__":
    main()
