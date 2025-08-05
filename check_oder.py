import os
import sys
import json
import time
import shutil
import warnings
import copy
from pathlib import Path
from datetime import datetime
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

import openpyxl

# Import Excel processor
from process_excel import process_excel_for_check_order

# Tắt warning openpyxl về default style
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

class OrderChecker:
    def ensure_template_excel(self):
        """
        Kiểm tra file input/template.xlsx, nếu chưa có thì tạo file mẫu với 2 cột: Tên viết tắt, Tên báo cáo
        """
        excel_path = self.input_dir / "template.xlsx"
        if not excel_path.exists():
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append(["Tên viết tắt", "Tên báo cáo"])
            ws.append(["DHTC", "DHTC - Đơn hàng thành công"])
            wb.save(excel_path)
    def load_report_list_from_excel(self):
        """
        Đọc danh sách báo cáo từ input/template.xlsx
        Trả về list các dict: { 'short_name': ..., 'report_name': ... }
        """
        excel_path = self.input_dir / "template.xlsx"
        if not excel_path.exists():
            return None
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            return None
        report_list = []
        for row in rows:
            if row[0] and row[1]:
                report_list.append({
                    'short_name': str(row[0]).strip(),
                    'report_name': str(row[1]).strip()
                })
        if not report_list:
            return None
        return report_list
    def __init__(self):
        self.base_path = self.get_base_path()
        self.input_dir = self.setup_directory("input")
        self.output_dir = self.setup_directory("output")
        self.daily_output_dir = self.setup_daily_output_directory()
        self.ensure_template_excel()
        self.chromium_path = self.get_chromium_path()
        self.config_file = self.input_dir / "config.json"
        self.config = self.load_or_create_config()
        
    def get_base_path(self):
        """
        Kiểm tra chế độ chạy và trả về đường dẫn base
        - Dev mode: thư mục chứa file .py
        - Packaged mode: thư mục chứa file .exe hoặc _internal
        """
        if getattr(sys, 'frozen', False):
            # Chế độ đóng gói (exe)
            # Kiểm tra xem có phải --onefile hay --onedir
            if hasattr(sys, '_MEIPASS'):
                # --onefile: sys._MEIPASS là thư mục tạm
                # Base path là thư mục chứa .exe
                base_path = Path(sys.executable).parent
            else:
                # --onedir: sys.executable là trong thư mục dist
                base_path = Path(sys.executable).parent
        else:
            # Chế độ dev
            base_path = Path(__file__).parent
            
        return base_path
    
    def setup_directory(self, dir_name):
        """
        Tạo thư mục nếu chưa tồn tại
        Trong chế độ --onedir, file được đặt trong _internal
        """
        if getattr(sys, 'frozen', False):
            # Packaged mode - tìm trong _internal trước
            internal_path = self.base_path / "_internal" / dir_name
            if internal_path.exists():
                return internal_path
            
        # Dev mode hoặc fallback
        dir_path = self.base_path / dir_name
        if not dir_path.exists():
            dir_path.mkdir(parents=True, exist_ok=True)
        return dir_path
    
    def get_chromium_path(self):
        """
        Tìm đường dẫn Chromium trong thư mục dự án
        """
        # Danh sách thư mục để tìm Chromium theo chế độ
        if getattr(sys, 'frozen', False):
            # Packaged mode - tìm trong _internal hoặc cùng cấp
            chromium_base_dirs = [
                self.base_path / "_internal" / "chromium-browser",  # --onedir
                self.base_path / "chromium-browser",               # Backup location
                self.base_path / "chromium",
            ]
        else:
            # Dev mode - tìm trong thư mục project
            chromium_base_dirs = [
                self.base_path / "chromium-browser",
                self.base_path / "chromium", 
                self.base_path / "browser",
            ]
        
        for base_dir in chromium_base_dirs:
            if base_dir.exists():
                # Tìm file chrome.exe trong tất cả thư mục con
                for chrome_exe in base_dir.rglob("chrome.exe"):
                    return str(chrome_exe)
        
        return None
    
    def install_browser(self):
        """
        Cài đặt Chromium vào thư mục dự án
        """
        browser_dir = self.base_path / "chromium-browser"
        
        # Tạo thư mục browser nếu chưa có
        browser_dir.mkdir(exist_ok=True)
        
        # Set environment variable để playwright cài vào thư mục này
        os.environ['PLAYWRIGHT_BROWSERS_PATH'] = str(browser_dir)
        
        # Cài đặt chromium
        os.system(f'"{sys.executable}" -m playwright install chromium')
        
        # Tìm lại đường dẫn sau khi cài
        self.chromium_path = self.get_chromium_path()
    
    def run_browser_test(self, url=None):
        """
        Chạy automation: đăng nhập, duyệt từng báo cáo trong template.xlsx, tải về và đặt tên file theo Tên viết tắt
        """
        if not self.config:
            return False

        report_list = self.load_report_list_from_excel()
        if not report_list:
            return False

        if not url:
            url = self.config['website']['url']

        try:
            with sync_playwright() as p:
                if self.chromium_path and os.path.exists(self.chromium_path):
                    browser = p.chromium.launch(
                        executable_path=self.chromium_path,
                        headless=True
                    )
                else:
                    browser = p.chromium.launch(headless=True)

                context = browser.new_context()
                page = context.new_page()

                login_success = self.login_to_website(page)
                if not login_success:
                    print("❌ Đăng nhập thất bại!")
                    return False

                navigation_success = self.navigate_to_reports(page)
                if not navigation_success:
                    print("❌ Điều hướng thất bại!")
                    return False

                # Lặp qua từng báo cáo trong Excel
                all_success = True
                downloaded_files = []
                
                print("─" * 60)
                print("📥 Đang tải báo cáo...")
                for idx, report in enumerate(report_list, 1):
                    success = self.select_kpi_and_download(page, report['report_name'], report['short_name'])
                    if success:
                        downloaded_files.append(report['short_name'])
                        print(f"✅ Tải file {idx}: {report['short_name']} thành công")
                    else:
                        print(f"❌ Tải file {idx}: {report['short_name']} thất bại")
                        all_success = False
                
                if downloaded_files:
                    print(f"✅ Đã tải thành công {len(downloaded_files)} báo cáo")
                print("─" * 60)
                
                # Đóng browser an toàn
                try:
                    context.close()
                    browser.close()
                except:
                    pass
                
                # Thông báo hoàn thành
                if downloaded_files:
                    print("\n" + "─" * 60)
                    print("📊 Đang xử lý file Excel...")
                    try:
                        process_success = process_excel_for_check_order()
                        if process_success:
                            print("✅ Xử lý Excel hoàn thành!")
                        else:
                            print("⚠️ Xử lý Excel có vấn đề!")
                    except Exception as e:
                        print(f"❌ Lỗi xử lý Excel: {str(e)}")
                    print(f"📁 Các file đã được lưu tại: {self.daily_output_dir}")
                    print("─" * 60)
                
                return all_success
        except Exception as e:
            print(f"❌ Lỗi browser: {str(e)}")
            return False

    def select_kpi_and_download(self, page, kpi_text, short_name):
        """
        Chọn KPI theo tên (kpi_text), tải file và đặt tên theo short_name
        """
        try:
            kpi_dropdown = self.get_locator(page, self.config['selectors']['kpi_dropdown'])
            kpi_dropdown.click()
            try:
                kpi_dropdown.select_option(label=kpi_text)
            except Exception as e:
                try:
                    dhtc_option = self.get_locator(page, f'//option[contains(text(), "{kpi_text}")]')
                    dhtc_option.click()
                except Exception as e2:
                    try:
                        dhtc_option = self.get_locator(page, f'//option[contains(text(), "{kpi_text.split()[0]}")]')
                        dhtc_option.click()
                    except Exception as e3:
                        return False

            selected_text = kpi_dropdown.locator('option:checked').inner_text()
            if kpi_text in selected_text or kpi_text.split()[0] in selected_text:
                return self.click_search_and_download(page, custom_filename=short_name)
            else:
                return False
        except PlaywrightTimeoutError:
            return False
        except Exception as e:
            return False

    def load_or_create_config(self):
        """
        Tải hoặc tạo file config.json
        """
        if not self.config_file.exists():
            print("📝 Tạo file config.json từ template...")
            self.create_default_config()
            # Thử load lại sau khi tạo
            try:
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                return config
            except:
                return None
        
        try:
            with open(self.config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            # Kiểm tra thông tin bắt buộc
            if not config.get('credentials', {}).get('username') or not config.get('credentials', {}).get('password'):
                print("❌ Config thiếu username/password!")
                return None
            
            if not config.get('website', {}).get('url'):
                print("❌ Config thiếu URL website!")
                return None
            
            return config
            
        except json.JSONDecodeError as e:
            print(f"❌ Config JSON không hợp lệ: {e}")
            return None
        except Exception as e:
            print(f"❌ Lỗi đọc config: {e}")
            return None
    
    def create_default_config(self):
        """
        Tạo file config từ template
        """
        template_file = self.input_dir / "config.template.json"
        
        # Kiểm tra file template có tồn tại không
        if template_file.exists():
            try:
                # Copy từ template
                with open(template_file, 'r', encoding='utf-8') as f:
                    template_config = json.load(f)
                
                # Lưu thành config.json
                with open(self.config_file, 'w', encoding='utf-8') as f:
                    json.dump(template_config, f, indent=2, ensure_ascii=False)
                
                print(f"✅ Đã tạo config.json từ template")
                return
                
            except Exception as e:
                print(f"❌ Lỗi đọc template: {e}")
        
        # Fallback: tạo config mặc định nếu không có template
        default_config = {
            "website": {
                "url": "https://example.com/login",
                "name": "Order Management System"
            },
            "credentials": {
                "username": "",
                "password": ""
            },
            "selectors": {
                "username_field": "#username",
                "password_field": "#password", 
                "login_button": "#login-button",
                "dashboard_indicator": ".dashboard",
                "search_button": "#search-btn",
                "menu_638": "#menu-638",
                "dms_report_kpi": "#dms-report-kpi",
                "rpt_kpi_staff": "#rpt-kpi-staff",
                "kpi_dropdown": "#kpi-dropdown",
                "dhtc_option_text": "DHTC - Đơn hàng thành công",
                "from_month_field": "#from-month-field",
                "month_year_picker_year": "#month-year-picker-year",
                "month_year_picker_month": "//select[@id='month-year-picker-month']/option[text()='T{month}']"
            },
            "settings": {
                "wait_timeout": 30000,
                "auto_wait": True,
                "fast_mode": False
            }
        }
        
        with open(self.config_file, 'w', encoding='utf-8') as f:
            json.dump(default_config, f, indent=2, ensure_ascii=False)
        
        print(f"✅ Đã tạo config.json mặc định")
    
    def login_to_website(self, page):
        """
        Đăng nhập vào website sử dụng thông tin từ config
        """
        if not self.config:
            return False
            
        try:
            url = self.config['website']['url']
            username = self.config['credentials']['username']
            password = self.config['credentials']['password']
            selectors = self.config['selectors']
            timeout = self.config['settings']['wait_timeout']
            
            # Mở trang đăng nhập
            page.goto(url)
            page.wait_for_load_state('networkidle')
            
            # Tìm và điền username
            try:
                username_field = self.get_locator(page, selectors['username_field'])
                self.smart_wait_for_element(username_field, 'visible')
                username_field.fill(username)
            except PlaywrightTimeoutError:
                return False
            
            # Tìm và điền password
            try:
                password_field = self.get_locator(page, selectors['password_field'])
                self.smart_wait_for_element(password_field, 'visible')
                password_field.fill(password)
            except PlaywrightTimeoutError:
                return False
            
            # Click nút đăng nhập
            try:
                login_button = self.get_locator(page, selectors['login_button'])
                self.smart_wait_for_element(login_button, 'visible')
                login_button.click()
            except PlaywrightTimeoutError:
                return False
            
            # Đợi trang dashboard load
            try:
                dashboard_indicator = self.get_locator(page, selectors['dashboard_indicator'])
                self.smart_wait_for_element(dashboard_indicator, 'visible')
                return True
                
            except PlaywrightTimeoutError:
                return False
                
        except Exception as e:
            return False

    def navigate_to_reports(self, page):
        """
        Điều hướng đến trang báo cáo sau khi đăng nhập thành công
        """
        try:
            menu_item_638 = self.get_locator(page, self.config['selectors']['menu_638'])
            menu_item_638.click()
            page.wait_for_load_state('networkidle')

            dms_report_kpi = self.get_locator(page, self.config['selectors']['dms_report_kpi'])
            dms_report_kpi.click()

            rpt_kpi_staff = self.get_locator(page, self.config['selectors']['rpt_kpi_staff'])
            try:
                rpt_kpi_staff.scroll_into_view_if_needed()
            except:
                pass
            rpt_kpi_staff.click()
            page.wait_for_load_state('networkidle')
            return True
        except PlaywrightTimeoutError:
            return False
        except Exception as e:
            return False

    def select_kpi_dropdown(self, page):
        """
        Chọn KPI từ dropdown lstKPI
        """
        try:
            print("   Looking for dropdown #lstKPI...")
            kpi_dropdown = self.get_locator(page, self.config['selectors']['kpi_dropdown'])
            kpi_dropdown.click()
            print("✅ Step 3a: Clicked KPI dropdown")

            print(f"   Selecting '{self.config['selectors']['dhtc_option_text']}'...")
            try:
                kpi_dropdown.select_option(label=self.config['selectors']['dhtc_option_text'])
                print("✅ Step 3b: Selected DHTC option using label")
            except Exception as e:
                print(f"   Method 1 failed: {e}")
                try:
                    option_text = self.config['selectors']['dhtc_option_text']
                    dhtc_option = self.get_locator(page, f'//option[contains(text(), "{option_text}")]')
                    dhtc_option.click()
                    print("✅ Step 3b: Selected DHTC option using text search")
                except Exception as e2:
                    print(f"   Method 2 failed: {e2}")
                    try:
                        dhtc_option = self.get_locator(page, '//option[contains(text(), "DHTC")]')
                        dhtc_option.click()
                        print("✅ Step 3b: Selected DHTC option using partial text")
                    except Exception as e3:
                        print(f"   All methods failed: {e3}")
                        return False

            selected_text = kpi_dropdown.locator('option:checked').inner_text()
            expected_text = self.config['selectors']['dhtc_option_text']
            if expected_text in selected_text or "DHTC" in selected_text:
                print(f"✅ Step 3c: Confirmed DHTC option is selected: {selected_text}")
                print("🔍 Step 4: Looking for Search button...")
                search_success = self.click_search_and_download(page)
                if search_success:
                    print("✅ Step 4: Search and download completed successfully!")
                    return True
                else:
                    print("❌ Step 4: Search and download failed!")
                    return False
            else:
                print(f"❌ Step 3c: Wrong option selected. Current text: {selected_text}")
                return False
        except PlaywrightTimeoutError:
            print("❌ Step 3 failed: KPI dropdown not found")
            return False
        except Exception as e:
            print(f"❌ KPI selection error: {e}")
            return False

    def select_month_year_before_search(self, page):
        """
        Chọn tháng/năm trước khi click search
        Chọn ngày hiện tại lùi 1 ngày để xác định tháng/năm
        """
        try:
            from datetime import datetime, timedelta
            
            # Tính ngày hiện tại lùi 1 ngày
            yesterday = datetime.now() - timedelta(days=1)
            target_year = yesterday.year
            target_month = yesterday.month
            
            print(f"📅 Chọn tháng/năm: {target_month}/{target_year}")
            
            # Bước 1: Click vào field fromMonth để mở month/year picker
            from_month_field = self.get_locator(page, self.config['selectors']['from_month_field'])
            from_month_field.click()
            page.wait_for_timeout(1000)  # Đợi picker hiện lên
            
            # Bước 2: Chọn năm
            year_selector = self.get_locator(page, self.config['selectors']['month_year_picker_year'])
            year_selector.select_option(value=str(target_year))
            print(f"✅ Đã chọn năm: {target_year}")
            page.wait_for_timeout(500)
            
            # Bước 3: Chọn tháng
            month_selector_template = self.config['selectors']['month_year_picker_month']
            month_selector = month_selector_template.format(month=target_month)
            month_element = self.get_locator(page, month_selector)
            month_element.click()
            print(f"✅ Đã chọn tháng: T{target_month}")
            page.wait_for_timeout(1000)
            
            return True
            
        except Exception as e:
            print(f"❌ Lỗi chọn tháng/năm: {str(e)}")
            return False

    def click_search_and_download(self, page, custom_filename=None):
        """
        Click nút Search và xử lý download file với retry logic
        Nếu custom_filename được truyền vào thì đặt tên file tải về theo tên này
        """
        max_retries = 3
        for attempt in range(1, max_retries + 1):
            try:
                # Bước mới: Chọn tháng/năm trước khi search
                month_year_success = self.select_month_year_before_search(page)
                if not month_year_success:
                    print(f"   ❌ Không thể chọn tháng/năm, thử lại lần {attempt}")
                    if attempt < max_retries:
                        if self.retry_navigation_from_menu(page):
                            continue
                    return False
                
                search_button = self.get_locator(page, self.config['selectors']['search_button'])
                if custom_filename:
                    download_path = self.daily_output_dir / f"{custom_filename}.xlsx"
                else:
                    download_path = self.daily_output_dir / f"report_{int(time.time())}.xlsx"
                with page.expect_download(timeout=180000) as download_info:
                    search_button.click()
                download = download_info.value
                download.save_as(str(download_path))
                if download_path.exists() and download_path.stat().st_size > 0:
                    return True
                else:
                    if attempt < max_retries:
                        if self.retry_navigation_from_menu(page):
                            continue
                    print(f"❌ Tải thất bại: {custom_filename or download_path.name}")
                    return False
            except Exception as e:
                if attempt < max_retries:
                    if self.retry_navigation_from_menu(page):
                        continue
                else:
                    print(f"❌ Tải thất bại: {custom_filename or 'Unknown'} - Lỗi: {str(e)}")
                    return False
        return False
    
    def retry_navigation_from_menu(self, page):
        """
        Thực hiện lại navigation từ menu #638 khi retry
        """
        try:
            timeout = self.config['settings']['wait_timeout']
            
            # Refresh page và đợi load
            page.reload()
            page.wait_for_load_state('networkidle')
            page.wait_for_timeout(3000)
            
            # Click menu #638
            menu_item_638 = self.get_locator(page, self.config['selectors']['menu_638'])
            menu_item_638.wait_for(state='visible', timeout=timeout)
            menu_item_638.click()
            
            page.wait_for_load_state('networkidle')
            page.wait_for_timeout(2000)
            
            # Click DMS_REPORT_KPI
            dms_report_kpi = self.get_locator(page, self.config['selectors']['dms_report_kpi'])
            dms_report_kpi.wait_for(state='visible', timeout=timeout)
            dms_report_kpi.click()
            
            page.wait_for_timeout(1000)
            
            # Click RPT_KPI_STAFF
            rpt_kpi_staff = self.get_locator(page, self.config['selectors']['rpt_kpi_staff'])
            rpt_kpi_staff.wait_for(state='visible', timeout=timeout)
            rpt_kpi_staff.click()
            
            page.wait_for_load_state('networkidle')
            page.wait_for_timeout(2000)
            
            # Bỏ việc chọn KPI dropdown ở đây vì sẽ được xử lý trong select_kpi_and_download
            
            return True
            
        except Exception as e:
            return False

    def smart_wait_for_element(self, locator, state='visible'):
        """
        Smart waiting strategy - sử dụng auto wait hoặc custom timeout
        """
        settings = self.config.get('settings', {})
        auto_wait = settings.get('auto_wait', True)
        fast_mode = settings.get('fast_mode', False)
        
        if auto_wait:
            # Sử dụng Playwright default timeout (30s) - nhanh và thông minh
            if fast_mode:
                # Fast mode: timeout ngắn hơn
                locator.wait_for(state=state, timeout=5000)  # 5 giây
            else:
                # Normal mode: dùng default của Playwright
                locator.wait_for(state=state)  # 30 giây default
        else:
            # Custom timeout từ config
            timeout = settings.get('wait_timeout', 30000)
            locator.wait_for(state=state, timeout=timeout)

    def get_locator(self, page, selector):
        """
        Tạo locator từ selector (hỗ trợ CSS và XPath)
        """
        if selector.startswith('/'):
            # XPath selector
            return page.locator(f"xpath={selector}")
        else:
            # CSS selector
            return page.locator(selector)

    def setup_daily_output_directory(self):
        """
        Tạo thư mục output theo ngày hiện tại (format: DDMMYYYY)
        Nếu đã tồn tại thì chỉ xóa các file Excel cũ, không xóa cả thư mục
        """
        today = datetime.now().strftime("%d%m%Y")
        daily_dir = self.output_dir / today
        
        if daily_dir.exists():
            # Chỉ xóa các file Excel cũ, không xóa cả thư mục
            try:
                for excel_file in daily_dir.glob("*.xlsx"):
                    try:
                        excel_file.unlink()  # Xóa file
                    except PermissionError:
                        print(f"⚠️ File {excel_file.name} đang được mở, bỏ qua...")
            except Exception as e:
                print(f"⚠️ Không thể dọn dẹp file cũ: {e}")
        else:
            daily_dir.mkdir(parents=True, exist_ok=True)
        
        return daily_dir

def main():
    """
    Hàm main chạy chương trình
    """
    # Khởi tạo OrderChecker
    checker = OrderChecker()
    
    # Chạy automation
    success = checker.run_browser_test()
    
    if success:
        print("✅ Hoàn thành tất cả!")
    else:
        print("❌ Có lỗi xảy ra!")

if __name__ == "__main__":
    main()